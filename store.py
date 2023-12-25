import datetime
from customtkinter import *
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Alignment

set_appearance_mode("light")
set_default_color_theme("green")

app = CTk()
app.geometry("1200x540+50+50")
app.title("2-1 Cafe")

wb = load_workbook('account.xlsx')
ws = wb.active
cafe_ws = wb["cafe '온점' 장부"]

visit_row = 4
visit_dic = {}
total_visit = 0

while cafe_ws.cell(column=11, row=visit_row).value != "총 방문자 수":
    visit_dic[str(cafe_ws.cell(column=11, row=visit_row).value)] = int(cafe_ws.cell(column=12, row=visit_row).value)
    visit_row += 1

try:
    list(visit_dic.keys()).index(str(datetime.datetime.now().date()).replace('-', '.'))
    total_visit = cafe_ws.cell(column=12, row=visit_row).value
except:
    cafe_ws.cell(column=11, row=visit_row).value = str(datetime.datetime.now().date()).replace('-', '.')
    cafe_ws.cell(column=12, row=visit_row).value = 0
    cafe_ws.cell(column=11, row=visit_row + 1).value = "총 방문자 수"
    total_visit = sum(visit_dic.values())
    cafe_ws.cell(column=12, row=visit_row + 1).value = total_visit
    visit_row += 1

if total_visit != 0:
    ws.row_dimensions[total_visit+3].height = 17
    cafe_ws.cell(column=2, row=total_visit+4).value = ''
    cafe_ws.cell(column=4, row=total_visit+4).value = ''
    cafe_ws.cell(column=5, row=total_visit+4).value = ''
    cafe_ws.cell(column=6, row=total_visit+4).value = ''
    cafe_ws.cell(column=7, row=total_visit+4).value = ''
    cafe_ws.cell(column=9, row=total_visit+4).value = ''

iced_tea_n = 0
misu_n = 0
watermelon_n = 0
toast_n = 0
total_price = 0

total_iced_tea = 0
total_misu = 0
total_watermelon = 0
total_toast = 0

cafe_menu = {
    'iced_tea' : 1000,
    'misu' : 1000,
    'watermelon' : 2000,
    'toast' : 1000
}

def act(act):
    if act == "add":
        return 1
    elif act == "delete":
        return -1
    
def count(var, action):
    global iced_tea_n, misu_n, watermelon_n, toast_n, total_price
    if var == 'iced_tea':
        iced_tea_n += act(action)
        iced_tea_Label.configure(text=f'아이스티: {iced_tea_n}개')
    elif var == 'misu':
        misu_n += act(action)
        misu_Label.configure(text=f'미숫가루: {misu_n}개')
    elif var == 'watermelon':
        watermelon_n += act(action)
        watermelon_Label.configure(text=f'수박 화채: {watermelon_n}개')
    elif var == 'toast':
        toast_n += act(action)
        toast_Label.configure(text=f'딸기잼 토스트: {toast_n}개')

    if iced_tea_n < 0:
        iced_tea_n = 0
        iced_tea_Label.configure(text=f'아이스티: {iced_tea_n}개')
    if misu_n < 0:
        misu_n = 0
        misu_Label.configure(text=f'미숫가루: {misu_n}개')
    if watermelon_n < 0:
        watermelon_n = 0
        watermelon_Label.configure(text=f'수박 화채: {watermelon_n}개')
    if toast_n < 0:
        toast_n = 0
        toast_Label.configure(text=f'딸기잼 토스트: {toast_n}개')
    
    total_price = cafe_menu['iced_tea'] * iced_tea_n + cafe_menu['misu'] * misu_n + cafe_menu['watermelon'] * watermelon_n + cafe_menu['toast'] * toast_n
    if total_price == 0:
        Total_label.configure(text='0원')
    else:
        Total_label.configure(text=f'{str(total_price)[0:-3]},{str(total_price)[-3:]}원')

def write():
    global total_visit, iced_tea_n, misu_n, watermelon_n, toast_n, total_price, total_iced_tea, total_misu, total_watermelon, total_toast
    if total_price != 0:
        total_visit = cafe_ws.cell(column=12, row=visit_row).value + 1
        cafe_ws.cell(column=2, row=total_visit + 2).value = total_visit
        cafe_ws.cell(column=4, row=total_visit + 2).value = iced_tea_n
        cafe_ws.cell(column=5, row=total_visit + 2).value = misu_n
        cafe_ws.cell(column=6, row=total_visit + 2).value = watermelon_n
        cafe_ws.cell(column=7, row=total_visit + 2).value = toast_n
        cafe_ws.cell(column=9, row=total_visit + 2).value = total_price

        total_iced_tea += iced_tea_n
        total_misu += misu_n
        total_watermelon += watermelon_n
        total_toast += toast_n

        iced_tea_left.configure(text=f'총 {total_iced_tea}개 판매')
        misu_left.configure(text=f'총 {total_misu}개 판매')
        watermelon_left.configure(text=f'총 {total_watermelon}개 판매')
        toast_left.configure(text=f'총 {total_toast}개 판매')

        iced_tea_n = 0
        misu_n = 0
        watermelon_n = 0
        toast_n = 0
        total_price = 0

        iced_tea_Label.configure(text=f'아이스티: {iced_tea_n}개')
        misu_Label.configure(text=f'미숫가루: {misu_n}개')
        watermelon_Label.configure(text=f'수박 화채: {watermelon_n}개')
        toast_Label.configure(text=f'딸기잼 토스트: {toast_n}개')
        Total_label.configure(text='0원')

        cafe_ws.cell(column=12, row=visit_row).value = total_visit
        cafe_ws.cell(column=12, row=visit_row-1).value = int(cafe_ws.cell(column=12, row=visit_row-1).value) + 1

iced_tea_photo = CTkImage(light_image=Image.open(fp='Image/iced_tea.jpg'), size=(120, 120))
iced_tea_Btn = CTkButton(app, width=250, height=250, text='아이스티\n{}원'.format(cafe_menu['iced_tea']), font=('배달의민족 한나체 Pro', 30), image=iced_tea_photo, compound='top', command=lambda:[count('iced_tea', 'add')])
iced_tea_Btn.place(x=10, y=10)

misu_photo = CTkImage(light_image=Image.open(fp='Image/misu.jpg'), size=(120, 120))
misu_Btn = CTkButton(app, width=250, height=250, text='미숫가루\n{}원'.format(cafe_menu['misu']), font=('배달의민족 한나체 Pro', 30), image=misu_photo, compound='top', command=lambda:[count('misu', 'add')])
misu_Btn.place(x=280, y=10)

watermelon_photo = CTkImage(light_image=Image.open(fp='Image/watermelon.jpg'), size=(120, 120))
watermelon_Btn = CTkButton(app, width=250, height=250, text='수박 화채\n{}원'.format(cafe_menu['watermelon']), font=('배달의민족 한나체 Pro', 30), image=watermelon_photo, compound='top', command=lambda:[count('watermelon', 'add')])
watermelon_Btn.place(x=10, y=280)

toast_photo = CTkImage(light_image=Image.open(fp='Image/toast.jpg'), size=(120, 120))
toast_Btn = CTkButton(app, width=250, height=250, text='딸기잼 토스트\n{}원'.format(cafe_menu['toast']), font=('배달의민족 한나체 Pro', 30), image=toast_photo, compound='top', command=lambda:[count('toast', 'add')])
toast_Btn.place(x=280, y=280)

cal_Frame = CTkFrame(app, width=300, height=520, bg_color='white')
cal_Frame.place(x=550, y=10)

iced_tea_Label = CTkLabel(app, text='아이스티: 0개', font=('배달의민족 한나체 Pro', 30), bg_color='#DBDBDB')
misu_Label = CTkLabel(app, text='미숫가루: 0개', font=('배달의민족 한나체 Pro', 30), bg_color='#DBDBDB')
watermelon_Label = CTkLabel(app, text='수박 화채: 0개', font=('배달의민족 한나체 Pro', 30), bg_color='#DBDBDB')
toast_Label = CTkLabel(app, text='딸기잼 토스트: 0개', font=('배달의민족 한나체 Pro', 30), bg_color='#DBDBDB')
iced_tea_Label.place(x=570, y=50)
misu_Label.place(x=570, y=150)
watermelon_Label.place(x=570, y=250)
toast_Label.place(x=570, y=350)

Total_label = CTkLabel(app, text= '0원', font=('배달의민족 한나체 Pro', 50), bg_color='#DBDBDB')
Total_label.place(x=570, y=450)

iced_tea_minus = CTkButton(app, width=140, height=80, text='-', font=(('배달의민족 한나체 Pro', 50)), command=lambda:[count('iced_tea', 'delete')])
iced_tea_minus.place(x=870, y=20)

misu_minus = CTkButton(app, width=140, height=80, text='-', font=(('배달의민족 한나체 Pro', 50)), command=lambda:[count('misu', 'delete')])
misu_minus.place(x=870, y=120)

watermelon_minus = CTkButton(app, width=140, height=80, text='-', font=(('배달의민족 한나체 Pro', 50)), command=lambda:[count('watermelon', 'delete')])
watermelon_minus.place(x=870, y=220)

toast_minus = CTkButton(app, width=140, height=80, text='-', font=(('배달의민족 한나체 Pro', 50)), command=lambda:[count('toast', 'delete')])
toast_minus.place(x=870, y=320)

Count_Btn = CTkButton(app, width=300, height=100, text='계산 완료', font=(('배달의민족 한나체 Pro', 50)), fg_color="black", command=write)
Count_Btn.place(x=870, y=420)

iced_tea_left = CTkLabel(app, width=140, height=80, text='총 0개 판매', font=(('배달의민족 한나체 Pro', 25)))
misu_left = CTkLabel(app, width=140, height=80, text='총 0개 판매', font=(('배달의민족 한나체 Pro', 25)))
watermelon_left = CTkLabel(app, width=140, height=80, text='총 0개 판매', font=(('배달의민족 한나체 Pro', 25)))
toast_left = CTkLabel(app, width=140, height=80, text='총 0개 판매', font=(('배달의민족 한나체 Pro', 25)))

iced_tea_left.place(x=1030, y=20)
misu_left.place(x=1030, y=120)
watermelon_left.place(x=1030, y=220)
toast_left.place(x=1030, y=320)

app.mainloop()


if total_visit != 0:
    ws.row_dimensions[total_visit + 3].height = 4
    cafe_ws.cell(column=2, row=total_visit+4).value = '합계'
    cafe_ws.cell(column=4, row=total_visit+4).value = f'=sum(D3:D{total_visit+2})'
    cafe_ws.cell(column=5, row=total_visit+4).value = f'=sum(E3:E{total_visit+2})'
    cafe_ws.cell(column=6, row=total_visit+4).value = f'=sum(F3:F{total_visit+2})'
    cafe_ws.cell(column=7, row=total_visit+4).value = f'=sum(G3:G{total_visit+2})'
    cafe_ws.cell(column=9, row=total_visit+4).value = f'=sum(I3:I{total_visit+2})'

for cell in cafe_ws["B"]:
    cell.alignment = Alignment(horizontal="center")

for cell in cafe_ws["D"]:
    cell.alignment = Alignment(horizontal="center")

for cell in cafe_ws["E"]:
    cell.alignment = Alignment(horizontal="center")

for cell in cafe_ws["F"]:
    cell.alignment = Alignment(horizontal="center")

for cell in cafe_ws["G"]:
    cell.alignment = Alignment(horizontal="center")

for cell in cafe_ws["I"]:
    cell.alignment = Alignment(horizontal="center")

for cell in cafe_ws["K"]:
    cell.alignment = Alignment(horizontal="center")

for cell in cafe_ws["L"]:
    cell.alignment = Alignment(horizontal="center")

wb.save("account.xlsx")
wb.close()